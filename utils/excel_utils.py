import openpyxl
import os

class XlsxReader:

    @staticmethod
    def get_test_data(testcaseName):
        data = []
        flag =False
        try:
            filepath=os.path.join(os.path.abspath(os.curdir), 
            'resources','input_data.xlsx')
            book = openpyxl.load_workbook(filepath)
            sheet = book.active
            for i in range(1, sheet.max_row+1):
                if sheet.cell(row=i, column=1).value == testcaseName:
                    flag=True
                    for j in range(2, sheet.max_column+1):
                        if sheet.cell(row=i, column=j).value == None:
                            break
                        data.append(sheet.cell(row=i, column=j).value)
                    break
        except:
            print("Unable to read file")
            raise ValueError("File not found or unable to read file")
        if flag == False:
            print("{} test case name not found in excel sheet, please provide correct testcase name".format(testcaseName))
        return data

    @staticmethod
    def get_test_data_based_on_sheet(sheet_name, test_case_name):
        flag =False
        data = []
        try:
            filepath=os.path.join(os.path.abspath(os.curdir), 
                'resources','input_data.xlsx')
            book = openpyxl.load_workbook(filepath)
            sheet = book[sheet_name]
            for i in range(1, sheet.max_row+1):
                if sheet.cell(row=i, column=1).value == test_case_name:
                    flag =True
                    for j in range(2, sheet.max_column+1):
                        data.append(sheet.cell(row=i, column=j).value)
                    break
        except:
            print("Unable to read excel file")
            raise ValueError("file not found or unable to read file")
        if flag == False:
            print("{} test case name not found in excel sheet, please provide correct testcase name".format(testcaseName))
        return data
    